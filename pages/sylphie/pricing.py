import streamlit as st
import pandas as pd
import openpyxl
import datetime as dt
import sys
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from functions.sheets import *
from sap_scripts.generate_script import *
from sap_scripts.run_script import run_sap_script
from pages.sylphie.sap_login import login_screen
from functions.vkp2_run import *

def alter_pricing():
    mecs = st.selectbox("Mecânica de Preço:", ['De/Por', 'De/MSRP', 'Cancelar/Prorrogar']) 
    if mecs == "De/Por":
        today = dt.datetime.today().strftime("%d/%m/%Y")
        file = st.file_uploader("Insira a planilha de alteração", type='xlsx')
        file_path = st.text_input("Insira o caminho da pasta onde você está trabalhando: ")
        username = st.text_input("Usuário SAP", type="password")
        password = st.text_input("Senha SAP", type="password")
        if st.button('Executar automação'):
            sheet = pd.read_excel(file, skiprows=1, engine='openpyxl')
            sheet_pyxl = openpyxl.load_workbook(file)
            pyxl_names = sheet_pyxl.sheetnames
            # st.write(pyxl_names)
            requested = sheet_pyxl['Solicitados'] #worksheet pricing changer
            st.write(requested)
            file_details = {"Filename": file.name, "FileType":file.type, "Size": file.size}
            st.write(file_details)
            # st.dataframe(sheet) # just for logging

            headers = [
                requested['A1'].value,
                requested['B1'].value, 
                requested['C1'].value,
                requested['D1'].value,
                requested['E1'].value,
                requested['F1'].value, 
                requested['G1'].value, 
                requested['H1'].value, 
                requested['I1'].value, 
                requested['J1'].value, 
                requested['K1'].value,
            ]

            # catch all needed columns
            st.write("max_rows -> ", requested.max_row)
            skus_vtex = getColumn_values(requested, column="A")
            skus = getColumn_values(requested, column=f"B")
            description = getColumn_values(requested, column="C")
            supplier = getColumn_values(requested, column="D")
            action_type = getColumn_values(requested, column="E")
            start_date = getColumn_values(requested, column="F")
            end_date = getColumn_values(requested, column="G")
            price_policy = getColumn_values(requested, column="H")
            de_prices = getColumn_values(requested, column="I")
            por_prices = getColumn_values(requested, column="J")
            arred_prices = getColumn_values(requested, column="K")
            # discount_column = ((por_prices[2] / de_prices[2]) - 1) #"L" it needs a loop to calculate formula
            # st.write(discount_column)
            # vkp2_script(file_path, username, password, transaction, skus, store, start_date, end_date, export_path, filename)
            # do vkp2 thing
            # return the prices on an array or list or something like it
            requested_data = list(zip(reversed(skus_vtex), reversed(skus), reversed(description), reversed(supplier), reversed(action_type),
                                      reversed(start_date), reversed(end_date), reversed(price_policy), reversed(de_prices),
                                      reversed(por_prices), reversed(arred_prices)))
            requested_dataframe = pd.DataFrame(requested_data, columns=headers)
            # with pd.ExcelWriter(f'planilhas/tomorrow{file.name}',
            #    mode='w') as writer:
            #    requested_dataframe.to_excel(writer, sheet_name='testtando')

            # next steps -> remove 1st column, 1st row from transformed.xlsx 
            # then -> create a function createDataframe by using getColumn_values
            # then -> use the data to vkp2 WORKING NOW
            # pegar todas colunas OK
            # ver se o dataframe funciona com todas OK
            # busca vkp2 OK
            # funcao vkp2 retorna path-like or file-like object OK (retornando dataframe)
            # getColumn_values para cada coluna em vkp2()
            # bater colunas de preço DE - fazer uma matriz
            # com todas as colunas, e manipular as posições
            # por fim, escrever o dataframe da vkp2 no arquivo result que terá duas abas
            # solicitados e vkp2
            htm_df = vkp2_runner(username=username, password=password, transaction='VKP2', skus=skus, store='1950', start_date=start_date[0], end_date=start_date[0], export_path=file_path, filename='SYLP1')

            # Apply the function to columns J and Q
            # htm_df['Montante'] = htm_df['Montante'].apply(clean_currency)
            # htm_df['Montante.1'] = htm_df['Montante.1'].apply(clean_currency)
            # htm_df['Montante'] = htm_df['Montante'].apply(insert_comma)
            # htm_df['Montante.1'] = htm_df['Montante.1'].apply(insert_comma)

            if 'VKP2' not in sheet_pyxl.sheetnames:
                sheet_vkp2 = sheet_pyxl.create_sheet('VKP2')
            for row in dataframe_to_rows(htm_df, index=False, header=True):
                sheet_pyxl['VKP2'].append(row)
            save_path = f"{file_path}\\{file.name}"
            sheet_pyxl.save(save_path)
            st.success(f"Arquivo salvo com sucesso em: {save_path}")
            st.write(htm_df)
            #sheet_pyxl['']
            # último check -> preço arred deve ser menor que preço de.
            # depois disso, essa etapa está pronta.
            # resta organizar os caminhos dos arquivos (onde serão salvos e manipulados)